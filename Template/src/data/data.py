from openpyxl import load_workbook

class data:
    def __init__(self, Path):
        self.path = Path
        self.wb = load_workbook(filename=self.path, read_only=True)
        self.sheet = self.wb.active  # Ou self.wb['NomDeLaFeuille']
        self.nbRouge = 0
        self.nbBleu = 0
        self.prenomsBleu = []
        self.nomsBleu = []
        self.prenomsRouge = []
        self.nomsRouge = []
        self.parsedata()

    def parsedata(self):
        self.nbRouge = self.sheet['B2'].value
        self.nbBleu = self.sheet['B3'].value

        # Peupler Rouge
        for row in self.sheet.iter_rows(min_row=8, min_col=2, max_col=3):
            prenom_cell, nom_cell = row
            if prenom_cell.value is None or nom_cell.value is None:
                break  # Arrête la boucle si une cellule est vide
            self.prenomsRouge.append(prenom_cell.value)
            self.nomsRouge.append(nom_cell.value)

        # Peupler Bleu
        for row in self.sheet.iter_rows(min_row=8, min_col=5, max_col=6):
            prenom_cell, nom_cell = row
            if prenom_cell.value is None or nom_cell.value is None:
                break  # Arrête la boucle si une cellule est vide
            self.prenomsBleu.append(prenom_cell.value)
            self.nomsBleu.append(nom_cell.value)

        # Fermer le fichier
        self.wb.close()

    def getnbBleu(self):
        return self.nbBleu

    def getnbRouge(self):
        return self.nbRouge

    def getParticipantBleu(self, index):
        return self.prenomsBleu[index], self.nomsBleu[index]

    def getParticipantRouge(self, index):
        return self.prenomsRouge[index], self.nomsRouge[index]
